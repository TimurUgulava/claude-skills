#!/usr/bin/env python3
"""
Собирает keys.xlsx из классифицированного JSON по 6 категориям.

Структура листа:
- Один лист "Лист1"
- Для каждой категории: заголовок с emoji, описание, шапка таблицы, строки-данные,
  пустая разделительная строка

Usage:
    python scripts/export_xlsx.py <input_keys.json> <output.xlsx>

Формат input_keys.json:
{
    "1": {
        "name": "Коммерческие",
        "emoji": "🛒",
        "description": "Запросы с коммерческим намерением, указывающие на готовность к покупке.",
        "keys": [
            {"query": "заказать продвижение", "relevance": 1.0, "recommendation": null},
            ...
        ]
    },
    "2": {...},
    ...
    "6": {...}
}
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    print("ОШИБКА: openpyxl не установлен. Запусти: pip install openpyxl")
    sys.exit(1)


CATEGORY_ORDER = ["1", "2", "3", "4", "5", "6"]

CATEGORY_DEFAULTS = {
    "1": {"emoji": "🛒", "name": "Коммерческие",
          "description": "Запросы с коммерческим намерением, указывающие на готовность к покупке.",
          "show_recommendation": False},
    "2": {"emoji": "📖", "name": "Информационные",
          "description": "Запросы, связанные с поиском информации и изучением темы.",
          "show_recommendation": False},
    "3": {"emoji": "🎯", "name": "Long-tail",
          "description": "Длинные специфические запросы с высокой конверсией.",
          "show_recommendation": True},
    "4": {"emoji": "🔍", "name": "SEO",
          "description": "Основные SEO-запросы для оптимизации страницы.",
          "show_recommendation": True},
    "5": {"emoji": "🧠", "name": "Семантические",
          "description": "Латентно-семантические ключевые слова для расширения контекста.",
          "show_recommendation": True},
    "6": {"emoji": "🔬", "name": "Верхнего уровня",
          "description": "Аналитические ключевые слова, выявленные AI.",
          "show_recommendation": True},
}


def export(keys_data: dict, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    header_font = Font(bold=True, size=12)
    category_font = Font(bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    category_fill = PatternFill("solid", fgColor="D9E1F2")

    current_row = 1

    for cat_id in CATEGORY_ORDER:
        if cat_id not in keys_data:
            continue

        cat = keys_data[cat_id]
        defaults = CATEGORY_DEFAULTS[cat_id]
        emoji = cat.get("emoji", defaults["emoji"])
        name = cat.get("name", defaults["name"])
        description = cat.get("description", defaults["description"])
        show_rec = defaults["show_recommendation"]
        keys = cat.get("keys", [])

        if not keys:
            continue

        full_name = f"{emoji} {name} ключевые слова"
        ws.cell(row=current_row, column=1, value=full_name).font = category_font
        ws.cell(row=current_row, column=1).fill = category_fill
        current_row += 1

        ws.cell(row=current_row, column=1, value=description)
        current_row += 1

        current_row += 1

        ws.cell(row=current_row, column=1, value="Ключевое слово").font = header_font
        ws.cell(row=current_row, column=2, value="Релевантность").font = header_font
        if show_rec:
            ws.cell(row=current_row, column=3, value="Рекомендация").font = header_font
        for col in (1, 2, 3 if show_rec else None):
            if col:
                ws.cell(row=current_row, column=col).border = thin_border
        current_row += 1

        for key in sorted(keys, key=lambda x: -x.get("relevance", 0)):
            ws.cell(row=current_row, column=1, value=key["query"])
            ws.cell(row=current_row, column=2, value=round(key.get("relevance", 0), 2))
            if show_rec:
                ws.cell(row=current_row, column=3, value=key.get("recommendation", "—"))
            current_row += 1

        current_row += 1

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 70

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__)
        return 1

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not input_path.exists():
        print(f"ОШИБКА: файл {input_path} не найден")
        return 1

    with input_path.open(encoding="utf-8") as f:
        keys_data = json.load(f)

    export(keys_data, output_path)
    print(f"✅ Файл сохранён: {output_path}")

    total_keys = sum(len(cat.get("keys", [])) for cat in keys_data.values() if isinstance(cat, dict))
    print(f"   Всего ключей: {total_keys}")
    for cat_id in CATEGORY_ORDER:
        if cat_id in keys_data:
            count = len(keys_data[cat_id].get("keys", []))
            name = keys_data[cat_id].get("name", CATEGORY_DEFAULTS[cat_id]["name"])
            emoji = keys_data[cat_id].get("emoji", CATEGORY_DEFAULTS[cat_id]["emoji"])
            print(f"   {emoji} {name}: {count}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
